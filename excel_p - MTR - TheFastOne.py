from openpyxl import load_workbook
from selenium import webdriver
from time import sleep
#匯入所需模組，excel、爬蟲和時間模組

wbPath = input("Please input the full path of file, including sub-filename:") #檔案位置，含副檔名
#wsColumn = int(input("Please input the location of column of part number(C = 3): ")) #參照的part number column 位置
wb = load_workbook(filename = wbPath)
ws = wb.active


def getPartNumbers():
    referPartNumber = 2 #參照檔案中，row的序數。因為第2個row資料起始位置，故設定2。
    saveUpdatedRow = 2 #寫入檔案中，row的序數。理由同上。
    referPoNumber = 2
    saveUpdatedPoRow = 2
    saved_column = 12 #寫入檔案中，column的序數。可變更。
    count2 = 2 #計數器。用於計算執行了多少次的迴圈。
    option = webdriver.ChromeOptions()
    option.add_argument("headless")
    option.add_argument("window-size=1920,1080") #因背景執行會不定失效，故加入這執行緒。
    chrome = webdriver.Chrome(chrome_options=option)
    objectPath = "https://www.fastenal.com/mtr"
    chrome.get(objectPath)
    while referPartNumber <= ws.max_row:
        partNumber = ws.cell(row = referPartNumber, column = 6).value #取值
        poNumber = ws.cell(row = referPoNumber, column = 8).value
        print(count2, partNumber, poNumber)
        count2 += 1
        chrome.find_element_by_xpath('//*[@id="partNumber"]').clear()
        chrome.find_element_by_xpath('//*[@id="controlNumber"]').clear()
        chrome.find_element_by_xpath('//*[@id="partNumber"]').send_keys(partNumber)
        chrome.find_element_by_xpath('//*[@id="controlNumber"]').send_keys(poNumber)
        button = chrome.find_element_by_xpath('//*[@id="mtrSearchForm"]/table/tbody/tr[3]/td/input')
        button.click()
        sleep(1.2) #等待網頁執行上述回傳資料的時間。原本預設3。
        try:
            if chrome.find_element_by_xpath('//*[@id="mtrResults"]/table/tbody/tr[1]/td/a'):
                url = chrome.find_element_by_xpath('//*[@id="mtrResults"]/table/tbody/tr[1]/td/a')
                geturl = url.get_attribute('href')
                ws.cell(row = saveUpdatedRow, column = saved_column, value = geturl)
            elif chrome.find_element_by_xpath('//*[@id="mtrResults"]/table[1]/tbody/tr[1]/td/a'):
                url = chrome.find_element_by_xpath('//*[@id="mtrResults"]/table[1]/tbody/tr[1]/td/a')
                geturl = url.get_attribute('href')
                ws.cell(row = saveUpdatedRow, column = saved_column, value = geturl)
            else:
                print("IT DOESN'T WORK!!!!!!!!!!")
                url = 'NA'
                ws.cell(row = saveUpdatedRow, column = saved_column, value = url)
        except:
            print("IT DOESN'T WORK!!!!!!!!!!")
            url = 'NA'
            ws.cell(row = saveUpdatedRow, column = saved_column, value = url)
        referPartNumber += 1
        referPoNumber += 1
        saveUpdatedRow += 1
        saveUpdatedPoRow += 1
    wb.save('The fast one.xlsx')
    chrome.close()


getPartNumbers()
