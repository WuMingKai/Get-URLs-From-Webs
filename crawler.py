from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.support.select import Select
from time import sleep


wbPath = input("Please input the full path of file, including sub-filename:")
wsColumn = int(input("Please input the column number of partnumber(ex: C column= 3): "))

wb = load_workbook(filename = wbPath)
ws = wb.active
def getPartNumbers():
    rowStartFrom = i
    i = 2
    columnStartFrom = k
    k = 2 # i 與 k ，是對應從公司內部抓下來的excle檔案，row和column的起始位置。
    columnSavedLocation = j
    j = 1 # 新檔案儲存url的column位置
    count2 = 1
    while i <= ws.max_row:
        partNumber = ws.cell(row = i, column = wsColumn).value
        #option = webdriver.ChromeOptions()
        #option.add_argument("headless")
        #chrome = webdriver.Chrome("path", chrome_options=option)
        chrome = webdriver.Chrome()
        objectPath = "target url" + partNumber #目標網頁的位址為 url + 貨號
        chrome.get(objectPath)
        print(count2)
        count2 += 1
        if chrome.find_element_by_xpath('//*[@id="default-print-file-tr"]/td[1]/a'):
            downLoadPrint = chrome.find_element_by_xpath('//*[@id="default-print-file-tr"]/td[1]/a')
            sleep(3)
            url = downLoadPrint.get_attribute('href')
            ws.cell(row = k, column = j, value = url)
        elif chrome.find_element_by_xpath('//*[@id="supplier-info-content"]/h4'):
            nothing = chrome.find_element_by_ptah('//*[@id="supplier-info-content"]/h4')
            url = nothing.get_attribute('class')
            ws.cell(row = k, column = j, value = url)
        i += 1
        k += 1
    wb.save('The Updated One.xlsx')
    chrome.close()


getPartNumbers()
